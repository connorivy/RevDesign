from shear_wall_classes import *
from openpyxl import load_workbook

def main():
    project = Project()

    path = "P:\\\\Structural\\2021\\2115 - Dallas\\2115.147 Toll Brothers Addison Road\\5 - Project Documents\B - Calculations\CI\Wood Calcs\Shear Walls\\"
    #load excel file
    workbook = load_workbook(filename=path + "toll_brothers_shear_wall_input.xlsx", data_only=True)

    ws = workbook['input']

    prev_shear_wall_letter = ws['A2'].value
    args = []
    blank_count = 0
    current_row = 0

    for row in ws.iter_rows(min_row=2, max_col=8, max_row=1500, values_only=True):
        current_row += 1
        # quit the for loop once three consecutive blanks have been found
        if row[0] == None:
            blank_count += 1
            if blank_count > 3:
                current_row -= 2
                break
            continue
        else:
            blank_count = 0
            
        if row[0] != prev_shear_wall_letter:
            write_values_to_excel(ws, project, args, current_row, prev_shear_wall_letter)
            args = []
            prev_shear_wall_letter = row[0]

        if row[6]:
            args.append(*[row[1:7]])
        else:
            args.append(*[row[1:6]])

    write_values_to_excel(ws, project, args, current_row, prev_shear_wall_letter)
    workbook.save(path + "shear_wall_output.xlsx")

def write_values_to_excel(ws, project, args, row, wall_location):
    num_rows = len(args)
    stacked_shear_wall = StackedShearWall(project, *args)

    for i in range(num_rows):
        schedule_entry = project.shear_wall_schedule.get_shear_wall(stacked_shear_wall.shear_force[i], max(0,stacked_shear_wall.chord_forces[i] / 1000), wall_location)
        ws.cell(row = row - num_rows + i, column = 13, value=stacked_shear_wall.shear_force[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 14, value=stacked_shear_wall.overturning_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 15, value=stacked_shear_wall.resisting_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 16, value=stacked_shear_wall.total_moment[i] / 1000)
        ws.cell(row = row - num_rows + i, column = 17, value=max(0,stacked_shear_wall.chord_forces[i] / 1000))
        try:
            ws.cell(row = row - num_rows + i, column = 18, value=schedule_entry.name)
        except:
            ws.cell(row = row - num_rows + i, column = 18, value=f'ERROR AT WALL {wall_location} - NO VALUE HIGH ENOUGH')


main()